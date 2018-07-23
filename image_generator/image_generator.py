from PIL import ImageFont, ImageDraw, Image
from openpyxl import load_workbook
import os
import json

# TODO: We can use 3 times bigger fonts and resize to have nicer (anti-aliased) texts
# TODO: Or rewrite in matplotlib

# Image generation specific constants
HEIGHTS = (250, 500)
BASE = Image.open("images/bg.jpg").convert("RGBA")
BASE_WIDTH, BASE_HEIGHT = BASE.width, BASE.height

# Excel pre-processing
wb = load_workbook(os.path.join(os.path.dirname(__file__), "excel/nemet_szavak_content8.xlsx"))
ws = wb.active

ids, german_words, hungarian_words, captions, dates = [[cell.value for cell in ws[col] if cell.value is not None][1:]
                                                       for col in ("A", "B", "C", "D", "E")]
german_words = [w.rstrip() for w in german_words]
hungarian_words = [w.rstrip() for w in hungarian_words]

unix_timestamps = [int(d.timestamp()) for d in dates]


def german_font(point_size):
    return ImageFont.truetype("fonts/Oswald-Bold.ttf", point_size)


def hungarian_font(point_size):
    return ImageFont.truetype("fonts/Montserrat-Regular.otf", point_size)


def image_generator(german_word, hungarian_word, output_file_number):

    txt = Image.new("RGBA", BASE.size, (255, 255, 255, 0))

    drawing_context = ImageDraw.Draw(txt)

    point_size_g = 92
    point_size_h = 76
    german_w = BASE_WIDTH
    hungarian_w = BASE_WIDTH
    while BASE_WIDTH - german_w <= 50 or BASE_WIDTH - hungarian_w <= 50:
        g_font = german_font(point_size_g)
        h_font = hungarian_font(point_size_h)
        # Get size of text
        german_w, german_h = drawing_context.textsize(german_word, g_font)
        # print(german_w)
        hungarian_w, hungarian_h = drawing_context.textsize(hungarian_word, h_font)
        if BASE_WIDTH - german_w <= 50:
            point_size_g -= 2
        if BASE_WIDTH - hungarian_w <= 50:
            point_size_h -= 2

    drawing_context.text((BASE.size[0] / 2 - german_w / 2, HEIGHTS[0]), german_word, (0, 0, 0, 255), g_font)
    drawing_context.text((BASE.size[0] / 2 - hungarian_w / 2, HEIGHTS[1]), hungarian_word, (0, 0, 0, 255), h_font)

    output = Image.alpha_composite(BASE, txt)

    output.save("outputs/" + str(output_file_number) + ".png")

# Image generation
for g_w, h_w, o_f_n in zip(german_words, hungarian_words, ids):
    image_generator(g_w, h_w, o_f_n)

# Schedules in JSON
with open("outputs/times.json", "w") as f:
    json.dump({k: [v1, v2] for k, v1, v2 in zip(ids, captions, unix_timestamps)}, f)
